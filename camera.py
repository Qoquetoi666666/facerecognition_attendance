import cv2
import os
import csv
from datetime import datetime
import openpyxl
from openpyxl import Workbook, load_workbook


class VideoCamera(object):
    def __init__(self):
        # 1. Khởi tạo Camera và AI
        self.video = cv2.VideoCapture(0)
        self.recognizer = cv2.face.LBPHFaceRecognizer_create()

        # Load bộ não đã train
        trainer_path = 'trainer/trainer.yml'
        if os.path.exists(trainer_path):
            self.recognizer.read(trainer_path)
            print("[INFO] Đã load file huấn luyện thành công.")
        else:
            print("[ERROR] Không tìm thấy file trainer.yml. Hãy chạy train_model.py trước!")

        self.faceCascade = cv2.CascadeClassifier("haarcascade/haarcascade_frontalface_default.xml")

        # 2. ĐỊNH NGHĨA DANH SÁCH TÊN THEO ID
        # Index 1 tương ứng User.1 (Huan), Index 2 tương ứng User.2 (Ham)
        self.names = ['None', 'Huan', 'Ham']

        # 3. Thiết lập file Excel
        self.excel_path = 'data/DiemDanh_TongHop.xlsx'
        self.ensure_excel_structure()

    def ensure_excel_structure(self):
        """Tạo thư mục data và file Excel nếu chưa có"""
        if not os.path.exists('data'):
            os.makedirs('data')
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = datetime.now().strftime('%m-%Y')
            ws.append(['Họ và Tên', 'Số buổi điểm danh', 'Ngày điểm danh gần nhất'])
            wb.save(self.excel_path)

    def log_attendance(self, name):
        """Logic: Tăng số buổi trong Excel và chống trùng lặp trong ngày"""
        now = datetime.now()
        month_year = now.strftime('%m-%Y')
        today_date = now.strftime('%d/%m/%Y')

        try:
            wb = load_workbook(self.excel_path)
            if month_year not in wb.sheetnames:
                ws = wb.create_sheet(month_year)
                ws.append(['Họ và Tên', 'Số buổi điểm danh', 'Ngày điểm danh gần nhất'])
            else:
                ws = wb[month_year]

            found = False
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == name:
                    found = True
                    last_date = ws.cell(row=row, column=3).value
                    # Chỉ cập nhật nếu hôm nay người này chưa được ghi nhận
                    if last_date != today_date:
                        current_count = ws.cell(row=row, column=2).value or 0
                        ws.cell(row=row, column=2).value = current_count + 1
                        ws.cell(row=row, column=3).value = today_date
                        print(f"[SUCCESS] Đã điểm danh cho {name}. Tổng buổi: {current_count + 1}")
                    break

            if not found:
                ws.append([name, 1, today_date])
                print(f"[SUCCESS] Điểm danh buổi đầu tiên cho {name}")

            wb.save(self.excel_path)
        except Exception as e:
            print(f"Lỗi ghi Excel: {e}")

    def get_frame(self):
        success, image = self.video.read()
        if not success:
            return None

        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        faces = self.faceCascade.detectMultiScale(gray, 1.2, 5)

        for (x, y, w, h) in faces:
            id, confidence = self.recognizer.predict(gray[y:y + h, x:x + w])

            # Kiểm tra độ tin cậy (Confidence < 75 là khá chính xác)
            if (confidence < 50):
                if id < len(self.names):
                    name = self.names[id]
                    self.log_attendance(name)
                else:
                    name = "Unknown"
                color = (0, 255, 0)  # Xanh lá
            else:
                name = "Unknown"
                color = (0, 0, 255)  # Đỏ

            # Vẽ khung chữ nhật và hiển thị tên
            cv2.rectangle(image, (x, y), (x + w, y + h), color, 2)
            cv2.putText(image, name, (x + 5, y - 5), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)
            # Hiển thị % khớp (tùy chọn)
            cv2.putText(image, f"{round(100 - confidence)}%", (x + 5, y + h - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.6,
                        (255, 255, 0), 1)

        ret, jpeg = cv2.imencode('.jpg', image)
        return jpeg.tobytes()

    def __del__(self):
        if self.video.isOpened():
            self.video.release()